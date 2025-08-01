import pandas as pd
import whois
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def check_domain_availability(domain_name):
    try:
        result = whois.whois(domain_name)
        domain_exists = result.domain_name is not None

        if domain_exists:
            status = "Taken"
            print(f"❌ Taken: {domain_name}")
        else:
            status = "Available"
            print(f"✅ Available: {domain_name}")
    except Exception:
        status = "Probably Available"
        print(f"✅ Probably Available: {domain_name}")
    return status

def autofit_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

def main():
    try:
        df = pd.read_csv("domains.csv")
    except FileNotFoundError:
        print("❗ Error: 'domains.csv' not found.")
        return

    if "domain" not in df.columns:
        print("❗ Error: CSV must have a column named 'domain'")
        return

    results = []

    for domain in df['domain']:
        domain = domain.strip()
        if domain:
            status = check_domain_availability(domain)
            results.append({
                "domain": domain,
                "availability_status": status
            })
            time.sleep(1)

    # Save to Excel
    output_df = pd.DataFrame(results)
    output_path = "domain_results_colored.xlsx"
    output_df.to_excel(output_path, index=False)

    # Style Excel
    wb = load_workbook(output_path)
    ws = wb.active

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        value = cell.value.lower()
        if "taken" in value:
            cell.fill = red_fill
        elif "available" in value:
            cell.fill = green_fill

    # Autofit column widths
    autofit_columns(ws)

    wb.save(output_path)
    print(f"\n✅ Colored and formatted Excel saved as '{output_path}'")

if __name__ == "__main__":
    main()
